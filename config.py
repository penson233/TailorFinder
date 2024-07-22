from openpyxl import Workbook
#fofa
FOFA_EMAIL = ""
FOFA_KEY = ""


#hunter
hunter_key=""
hunter_name=""


#天眼查
tyccookie=""
tyctoken=""

#securitytrails
securitykey=""

#https://hunter.io
hunteremailkey=""


##设置持续话收集邮箱

emailaccount = ""
emailpassword = ""
emailtarget=['']
mailhost=""


wb = Workbook()

wb.create_sheet("domain",0)
wb.create_sheet("company_structs",1)
wb.create_sheet("旗下app",2)
wb.create_sheet("email",3)
wb.create_sheet("subfinder", 4)
wb.create_sheet("fofa", 5)
wb.create_sheet("hunter",6)
wb.create_sheet("crt.sh",7)
wb.create_sheet("assetfinder",8)
wb.create_sheet("shuffledns",9)
wb.create_sheet("securitytrails", 10)
wb.create_sheet("rappidns",11)
wb.create_sheet("ip138",12)
wb.create_sheet("alldomain",13)



rootdomain=wb["domain"]
emailwb=wb["email"]
appwb=wb["旗下app"]
subfinderwb = wb["subfinder"]
fofawb=wb["fofa"]
hunterwb=wb["hunter"]
crtwb=wb["crt.sh"]
ctwb=wb["company_structs"]
assetfinderwb=wb["assetfinder"]
alldomainwb=wb["alldomain"]
shufflednswb=wb["shuffledns"]
securitytrailswb=wb["securitytrails"]
rappidnswb=wb["rappidns"]
ip138wb=wb["ip138"]

hunterwb.append(['IP', '端口', '域名', 'IP标签', 'url', '网站标题', '高危协议', '协议', '通讯协议', '网站状态码', '应用/组件', '操作系统', '备案单位', '备案号', '备案异常', '资产标签', '国家', '省份', '市区', '探查时间', 'Web资产', '运营商', '注册机构', '证书SHA256'])
fofawb.append(["domain", 'header_code', 'ip', 'port', 'webtitle'])
#fofa查询语法
fofasearch= "domain=\"{}\"||cert=\"{}\""

#hunter查询语法
huntersearch='domain.suffix="{}"'



#工具参数
#shuffledns
#linux
l_shuffledns="cat %s | xargs -I {} ./bin/shuffledns/shuffledns -d {} -w ./bin/shuffledns/domain.txt -r ./bin/shuffledns/resolvers.txt  >> %s/shuffledns.txt -m ./bin/massdns/massdns_linux"
#macos
m_shuffledns="cat %s | xargs -I {} ./bin/shuffledns/shuffledns_m -d {} -w ./bin/shuffledns/domain.txt -r ./bin/shuffledns/resolvers.txt  >> %s/shuffledns.txt -m ./bin/massdns/massdns_linux"


#assetfinder
#linux
l_assetfinder="cat %s | xargs -I {} ./bin/assetfinder/assetfinder {} >> %s/assetfinder.txt"
#macos
m_assetfinder="cat %s | xargs -I {} ./bin/assetfinder/assetfinder_m {} >> %s/assetfinder.txt"

webport=[8000,8080,8089,9000,9200,11211,27017,80,81,82,83,84,85,86,87,88,89,90,91,92,98,99,443,800,801,808,880,888,889,1000,1010,1080,1081,1082,1099,1118,1888,2008,2020,2100,2375,2379,3000,3008,3128,3505,5555,6080,6648,6868,7000,7001,7002,7003,7004,7005,7007,7008,7070,7071,7074,7078,7080,7088,7200,7680,7687,7688,7777,7890,8000,8001,8002,8003,8004,8006,8008,8009,8010,8011,8012,8016,8018,8020,8028,8030,8038,8042,8044,8046,8048,8053,8060,8069,8070,8081,8082,8083,8084,8085,8086,8087,8088,8089,8090,8091,8092,8093,8094,8095,8096,8097,8098,8099,8100,8101,8108,8118,8161,8172,8180,8181,8200,8222,8244,8258,8280,8288,8300,8360,8443,8448,8484,8800,8834,8838,8848,8858,8868,8879,8880,8881,8888,8899,8983,8989,9000,9001,9002,9008,9010,9043,9060,9080,9081,9082,9083,9084,9085,9086,9087,9088,9089,9090,9091,9092,9093,9094,9095,9096,9097,9098,9099,9100,9200,9443,9448,9800,9981,9986,9988,9998,9999,10000,10001,10002,10004,10008,10010,10250,12018,12443,14000,16080,18000,18001,18002,18004,18008,18080,18082,18088,18090,18098,19001,20000,20720,21000,21501,21502,28018,20880]

#subfinder
#linux
l_subfinder="./bin/subfinder/subfinder -dL {} -o {}/subfinder.txt"
#mac
m_subfinder="./bin/subfinder/subfinder_m -dL {} -o {}/subfinder.txt"

#端口扫描参数 预留格式化参数{} outpath
masscan="masscan -iL {}/alive_ip.txt -p 21,22,80,81,135,139,443,445,1433,1521,3306,5432,6379,7001,8000,8080,8089,9000,9200,11211,27017,82,83,84,85,86,87,88,89,90,91,92,98,99,443,800,801,808,880,888,889,1000,1010,1080,1081,1082,1099,1118,1888,2008,2020,2100,2375,2379,3000,3008,3128,3505,5555,6080,6648,6868,7000,7001,7002,7003,7004,7005,7007,7008,7070,7071,7074,7078,7080,7088,7200,7680,7687,7688,7777,7890,8000,8001,8002,8003,8004,8006,8008,8009,8010,8011,8012,8016,8018,8020,8028,8030,8038,8042,8044,8046,8048,8053,8060,8069,8070,8080,8081,8082,8083,8084,8085,8086,8087,8088,8089,8090,8091,8092,8093,8094,8095,8096,8097,8098,8099,8100,8101,8108,8118,8161,8172,8180,8181,8200,8222,8244,8258,8280,8288,8300,8360,8443,8448,8484,8800,8834,8838,8848,8858,8868,8879,8880,8881,8888,8899,8983,8989,9000,9001,9002,9008,9010,9043,9060,9080,9081,9082,9083,9084,9085,9086,9087,9088,9089,9090,9091,9092,9093,9094,9095,9096,9097,9098,9099,9100,9200,9443,9448,9800,9981,9986,9988,9998,9999,10000,10001,10002,10004,10008,10010,10250,12018,12443,14000,16080,18000,18001,18002,18004,18008,18080,18082,18088,18090,18098,19001,20000,20720,21000,21501,21502,28018,20880 --rate=1000 -oJ {}/alive_port.json"

#nmap扫描部分参数
nmap="-sV -n -Pn --open -sT -T4 --version-intensity 7"

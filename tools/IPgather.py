import openpyxl
import re
from collections import Counter
from tools import colorprint
import os
import aiodns
import asyncio
import sys
import socket
from tools.commons import readdomain
if sys.platform == "win32" or sys.platform == "win64":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

ip_domain={}
real_domain=[]
all_ip_repeat=[]



async def query(domain, resolver,outpath):
    try:
        domains=readdomain(outpath+"/domain")

        domain_re=re.compile('|'.join(domains).replace('\n',''))

        if domain_re.search(domain):
            find_ip = re.compile("(\d+\.\d+\.\d+\.\d+)\.")
            if not find_ip.search(domain):
                result = await resolver.query(domain, 'A')
                print(f"Resolved {domain} to {result}")

                ipreg = re.compile("\d+\.\d+\.\d+\.\d+")

                ipfind=ipreg.findall(str(result))[0]
                find_c = re.compile("(\d+\.\d+\.\d+)\.")


                if ipfind not in ip_domain:
                    ip_domain[ipfind] = [domain]

                    ip_find = find_c.findall(ipfind)

                    all_ip_repeat.append(ip_find[0])
                    real_domain.append(domain)
                else:
                    ip_domain[ipfind].append(domain)

                print(f"Resolved {domain} to {result}")
        else:
            print("not find")


    except Exception as e:
        print(domain)
        with open(outpath+'/error_domain','a+') as fe:
            fe.write(str(domain)+"\n")

async def DomainSearch(outpath,workbook):
    resolver = aiodns.DNSResolver()
    all_domain_wb = workbook["alldomain"]
    columns = list(all_domain_wb.columns)


    tasks = []
    for all in columns[0]:
        if all.value != "alldomain":

            tasks.append(query(all.value, resolver,outpath))

    await asyncio.gather(*tasks)

def is_domain_alive(domain):
    try:
        # 使用socket库尝试解析域名的IP地址
        socket.gethostbyname(domain)
        return True
    except (socket.gaierror, socket.timeout):
        return False

def DomainsAlive(workbook):

    all_domain_wb = workbook["alldomain"]
    alive_wb=workbook['alive_domain']
    alive_wb.append(["is_alive"])

    columns = list(all_domain_wb.columns)
    for domain in columns[0]:
        try:
            if is_domain_alive(domain.value):
                print(f"{domain.value} is alive")
                alive_wb.append([domain.value])
            else:
                print(f"{domain.value} is not alive")
        except Exception as e:
            print(e)
            continue

def readxls(outpath,name):
    colorprint.Red("[-]start to gather ip....")
    workbook = openpyxl.load_workbook(outpath+"/test.xlsx")

    workbook.create_sheet("fofa_ip", 14)
    workbook.create_sheet("hunter_ip", 15)
    workbook.create_sheet("all_ip",16)
    workbook.create_sheet("alive_domain",17)

    fofa_iplist = []
    hunter_iplist = []
    ip_fan_fofa = []
    ip_fan_hunter = []


    '''获取fofa收集到的ip'''
    fofa = workbook["fofa"]
    find_c = re.compile("(\d+\.\d+\.\d+)\.")
    columns = list(fofa.columns)
    if len(columns)>0:
        for i in columns[2]:
            if i.value != "ip":
                if i.value not in ip_fan_fofa:
                    try:
                        c = find_c.findall(i.value)[0]
                        fofa_iplist.append(c)
                        ip_fan_fofa.append(i.value)
                    except:
                        continue

    fofa_all = Counter(fofa_iplist).most_common()

    '''获取hunter收集到的ip'''
    hunter = workbook["hunter"]
    columns = list(hunter.columns)
    if len(columns)>0:
        for i in columns[0]:
            if i.value != "IP":
                if i.value not in ip_fan_hunter:
                    try:
                        c = find_c.findall(i.value)[0]
                        hunter_iplist.append(c)
                        ip_fan_hunter.append(i.value)
                    except:
                        continue

    hunter_all = Counter(hunter_iplist).most_common()





    fofa_ip = workbook["fofa_ip"]
    fofa_ip.append(['c段', '次数'])

    for all in fofa_all:
        fofa_ip.append([all[0] + ".0/24", all[1]])

    hunter_ip = workbook["hunter_ip"]
    hunter_ip.append(["c段", "次数"])

    for all in hunter_all:
        hunter_ip.append([all[0] + ".0/24", all[1]])




    DomainsAlive(workbook)



    asyncio.run(DomainSearch(outpath,workbook))



    all_ip_wb = workbook["all_ip"]
    all_ip_wb.append(['c段', '次数', '去掉泛解析域名'])


    all_ip_conut= Counter(all_ip_repeat).most_common()

    real_wb=[]
    for all_ip in all_ip_conut:
        real_wb.append([all_ip[0] + ".0/24", all_ip[1]])

    with open(outpath+'/realdomain.txt','w') as f:

        for i in range(len(real_domain)):
            if i < len(real_wb):
                real_wb[i].append(real_domain[i])
            else:
                real_wb.append(["","",real_domain[i]])
            f.write(real_domain[i]+'\n')

    for j in real_wb:
        all_ip_wb.append(j)

    for key in ip_domain:
        print(key, ":", ip_domain[key])
        all_ip_wb.append([key,'-------'.join(ip_domain[key])])


    workbook.save(outpath+"/"+name+".xlsx")
    os.system("rm -f "+outpath+"/test.xlsx")